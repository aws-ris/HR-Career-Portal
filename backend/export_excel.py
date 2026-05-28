from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
import psycopg2

def run():
    conn = psycopg2.connect("postgresql://postgres:postgres@localhost:5432/ris_db")
    cur = conn.cursor()
    cur.execute("""
        SELECT 
            c.id, c.full_name, c.dob, c.mobile_no, c.email, 
            t.position_applied, t.admin_department,
            l.about, l.google_scholar,
            s.class_x_percentage, s.class_xii_percentage
        FROM candidate_metadata c
        LEFT JOIN application_tracking t ON t.candidate_id = c.id
        LEFT JOIN candidate_links_about l ON l.candidate_id = c.id
        LEFT JOIN candidate_schooling s ON s.candidate_id = c.id;
    """)
    applicant_cols = [desc[0] for desc in cur.description]
    applicants = cur.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Applicants"

    # Define the top-level categories and their sub-columns
    sections = [
        ("Basic Info", '4F81BD', ['Full Name', 'DOB', 'Mobile Number', 'Email', 'Description', 'Position Applied', 'Admin Department']),
        ("Secondary Education", '76923C', ['Class X', 'Class XII']),
        ("Graduation Details", '9BBB59', ['University', 'Degree', 'Score Type', 'Score Value', 'Year of Passing'] * 3),
        ("Postgraduate Details", '8064A2', ['University', 'Degree', 'Score Type', 'Score Value', 'Year of Passing'] * 3),
        ("Doctorate Details", '4BACC6', ['University', 'Thesis', 'Score Type', 'Score Value', 'Year of Passing'] * 3),
        ("Publications Summary", 'F79646', ['Types']),
        ("Books", 'FFC000', ['Book 1 Title', 'Book 2 Title', 'Book 3 Title']),
        ("Chapters", '00B050', ['Chapter 1 Title', 'Chapter 1 Book', 'Chapter 2 Title', 'Chapter 2 Book', 'Chapter 3 Title', 'Chapter 3 Book']),
        ("Papers", '7030A0', ['Paper 1 Title', 'Paper 2 Title', 'Paper 3 Title']),
        ("Google Scholar", 'A6A6A6', ['Link']),
        ("Work Experience Details", 'C0504D', ['Company', 'Role', 'Start Date', 'End Date'] * 3),
        ("Database Info", '808080', ['Applicant ID'])
    ]

    col_idx = 1
    for category, color, subcols in sections:
        start_col = col_idx
        end_col = col_idx + len(subcols) - 1
        
        ws.cell(row=1, column=start_col, value=category)
        if end_col > start_col:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        
        cell = ws.cell(row=1, column=start_col)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

        for i, subcol in enumerate(subcols):
            c = ws.cell(row=2, column=start_col + i, value=subcol)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')

        col_idx = end_col + 1

    row_idx = 3
    for app in applicants:
        app_dict = dict(zip(applicant_cols, app))
        app_id = app_dict['id']

        row_data = []

        # 1. Basic Info
        row_data.extend([
            app_dict['full_name'], str(app_dict['dob']), 
            app_dict['mobile_no'], app_dict['email'], app_dict['about'] or '', 
            app_dict['position_applied'], app_dict['admin_department'] or 'None', 
            app_dict['class_x_percentage'] or 0.0, app_dict['class_xii_percentage'] or 0.0
        ])

        # 2. Graduations (Level = undergrad)
        cur.execute("SELECT university, degree_name, score_type, score_value, grad_year FROM candidate_higher_education WHERE candidate_id = %s AND level = 'undergrad' ORDER BY entry_order LIMIT 3;", (app_id,))
        bachelors = cur.fetchall()
        for i in range(3):
            if i < len(bachelors):
                row_data.extend([bachelors[i][0], bachelors[i][1] or '', bachelors[i][2], bachelors[i][3], bachelors[i][4] or ''])
            else:
                row_data.extend(['', '', '', '', ''])

        # 3. Postgrad (Level = postgrad)
        cur.execute("SELECT university, degree_name, score_type, score_value, grad_year FROM candidate_higher_education WHERE candidate_id = %s AND level = 'postgrad' ORDER BY entry_order LIMIT 3;", (app_id,))
        masters = cur.fetchall()
        for i in range(3):
            if i < len(masters):
                row_data.extend([masters[i][0], masters[i][1] or '', masters[i][2], masters[i][3], masters[i][4] or ''])
            else:
                row_data.extend(['', '', '', '', ''])

        # 4. Doctorate (Level = phd)
        cur.execute("SELECT university, degree_name, score_type, score_value, grad_year FROM candidate_higher_education WHERE candidate_id = %s AND level = 'phd' ORDER BY entry_order LIMIT 3;", (app_id,))
        docs = cur.fetchall()
        for i in range(3):
            if i < len(docs):
                row_data.extend([docs[i][0], docs[i][1] or '', docs[i][2], docs[i][3], docs[i][4] or ''])
            else:
                row_data.extend(['', '', '', '', ''])

        # Publications fetch
        cur.execute("SELECT pub_type, title, parent_book FROM candidate_publications WHERE candidate_id = %s ORDER BY entry_order;", (app_id,))
        all_pubs = cur.fetchall()
        
        # 5. Publication Summary
        pub_types = list(set([p[0] for p in all_pubs]))
        row_data.append(", ".join(pub_types) if pub_types else "None")

        # 6. Books
        books = [p for p in all_pubs if p[0] == 'book'][:3]
        for i in range(3):
            row_data.append(books[i][1] if i < len(books) else '')

        # 7. Chapters
        chapters = [p for p in all_pubs if p[0] == 'chapter'][:3]
        for i in range(3):
            if i < len(chapters):
                row_data.extend([chapters[i][1], chapters[i][2] or '']) # title, parent_book
            else:
                row_data.extend(['', ''])

        # 8. Papers
        papers = [p for p in all_pubs if p[0] == 'paper'][:3]
        for i in range(3):
            row_data.append(papers[i][1] if i < len(papers) else '')

        # 9. Google Scholar
        row_data.append(app_dict.get('google_scholar') or '')

        # 10. Work Experience
        cur.execute("SELECT company_name, role, start_date, end_date FROM candidate_work_experience WHERE candidate_id = %s ORDER BY entry_order LIMIT 3;", (app_id,))
        works = cur.fetchall()
        for i in range(3):
            if i < len(works):
                row_data.extend([works[i][0], works[i][1], str(works[i][2]), str(works[i][3] if works[i][3] else 'Present')])
            else:
                row_data.extend(['', '', '', ''])

        # 11. Database Info (ID moved to end)
        row_data.append(str(app_id))

        # Write the row
        for c_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=c_idx, value=val)
        row_idx += 1

    from openpyxl.utils import get_column_letter

    # Auto-adjust column widths
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                if cell.value:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = min(adjusted_width, 50) # Cap width

    ws.freeze_panes = 'A3' # Freeze the 2 header rows
    
    filename = "senior_official_export.xlsx"
    wb.save(filename)
    print(f"Export complete: {filename}")

if __name__ == '__main__':
    run()
