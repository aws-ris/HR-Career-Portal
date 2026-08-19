import openpyxl

def check_scores():
    print("=== STANDARDIZED EXPORT SCORES ===")
    wb = openpyxl.load_workbook("test_standardized_export.xlsx")
    ws = wb.active
    
    # Headers: ['Full Name', 'Class X %', 'Class XII %', 'Bachelors (UG)', 'Bachelors Score', 'Bachelors Year', 'Masters (PG)', 'Masters Score', 'Masters Year', 'Doctorate (PhD)', 'Doctorate Score', 'Doctorate Year', 'Total Exp (Yrs)', 'Latest Employment']
    print(f"{'Name':<20} | {'Class X':<8} | {'Class XII':<9} | {'Bach Score':<12} | {'Mast Score':<12} | {'PhD Score':<12}")
    print("-" * 80)
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        x = ws.cell(row=r, column=2).value
        xii = ws.cell(row=r, column=3).value
        ug_score = ws.cell(row=r, column=5).value
        pg_score = ws.cell(row=r, column=8).value
        phd_score = ws.cell(row=r, column=11).value
        print(f"{str(name):<20} | {str(x):<8} | {str(xii):<9} | {str(ug_score):<12} | {str(pg_score):<12} | {str(phd_score):<12}")
    wb.close()

    print("\n=== DETAILED EXPORT SCORES ===")
    wb = openpyxl.load_workbook("test_detailed_export.xlsx")
    ws = wb.active
    # In detailed export:
    # Col 1 (Full Name), Col 12 (Class X %), Col 13 (Class XII %), Col 16 (Graduation Score), Col 20 (Postgrad Score), Col 24 (PhD Score)
    print(f"{'Name':<20} | {'Class X':<8} | {'Class XII':<9} | {'Grad Score':<12} | {'PG Score':<12} | {'PhD Score':<12}")
    print("-" * 80)
    for r in range(2, min(ws.max_row + 1, 40)):
        name = ws.cell(row=r, column=1).value
        x = ws.cell(row=r, column=12).value
        xii = ws.cell(row=r, column=13).value
        ug_score = ws.cell(row=r, column=16).value
        pg_score = ws.cell(row=r, column=20).value
        phd_score = ws.cell(row=r, column=24).value
        if name or x or xii or ug_score or pg_score or phd_score:
            print(f"{str(name or ''):<20} | {str(x or ''):<8} | {str(xii or ''):<9} | {str(ug_score or ''):<12} | {str(pg_score or ''):<12} | {str(phd_score or ''):<12}")
    wb.close()

if __name__ == '__main__':
    check_scores()
