import requests
import openpyxl
import json
import sys

def main():
    # 1. Fetch the Job ID for DAKSHIN Research Assistant (Development Finance) from Vercel
    # We can get it by listing jobs if there's an endpoint, or we can use the candidates we just fixed.
    # The candidate ID for Riya Nair is "0cd37a45-20b2-4cbc-afc7-b66a081684e3".
    # Let's hit the server to export the job candidates. We need the job ID first.
    # Let's fetch all jobs from the API to find the DAKSHIN one.
    jobs_url = "https://hr-portal-ris.vercel.app/api/v1/jobs"
    print(f"Fetching jobs from {jobs_url}...")
    res = requests.get(jobs_url)
    if res.status_code != 200:
        print(f"Error fetching jobs: {res.status_code}")
        print(res.text)
        sys.exit(1)
        
    jobs = res.json()
    job_id = None
    for j in jobs:
        if "Research Assistant (Development Finance)" in j.get("title", ""):
            job_id = j.get("id")
            print(f"Found Job: {j.get('title')} -> ID: {job_id}")
            break
            
    if not job_id:
        # Fallback to search for Research Assistant
        for j in jobs:
            if "Research Assistant" in j.get("title", ""):
                job_id = j.get("id")
                print(f"Fallback Job: {j.get('title')} -> ID: {job_id}")
                break
                
    if not job_id:
        print("DAKSHIN Research Assistant job not found in job list!")
        sys.exit(1)
        
    # 2. Call the export endpoint for detailed export
    export_url = f"https://hr-portal-ris.vercel.app/api/v1/jobs/{job_id}/candidates/export"
    print(f"Downloading detailed Excel from Vercel: {export_url}")
    payload = {
        "filters": {},
        "format": "xlsx",
        "columns": ["contact", "highest_edu", "grad", "pg", "phd", "work", "books", "papers", "chapters"],
        "report_type": "detailed"
    }
    
    res = requests.post(export_url, json=payload)
    if res.status_code != 200:
        print(f"Failed to export detailed: {res.status_code}")
        print(res.text)
        sys.exit(1)
        
    with open("vercel_detailed_export.xlsx", "wb") as f:
        f.write(res.content)
    print("Detailed export saved as vercel_detailed_export.xlsx")
    
    # 3. Call the export endpoint for standardized export
    print(f"Downloading standardized Excel from Vercel: {export_url}")
    payload["report_type"] = "standardized"
    res = requests.post(export_url, json=payload)
    if res.status_code != 200:
        print(f"Failed to export standardized: {res.status_code}")
        print(res.text)
        sys.exit(1)
        
    with open("vercel_standardized_export.xlsx", "wb") as f:
        f.write(res.content)
    print("Standardized export saved as vercel_standardized_export.xlsx")
    
    # 4. Verify formatting
    print("\n=== VERIFYING STANDARDIZED EXPORT SCORES (VERCEL) ===")
    wb = openpyxl.load_workbook("vercel_standardized_export.xlsx")
    ws = wb.active
    print(f"{'Name':<20} | {'Class X':<8} | {'Class XII':<9} | {'Bach Score':<12} | {'Mast Score':<12} | {'PhD Score':<12}")
    print("-" * 80)
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        # Skip Viraal Saini from prints to focus on fixed candidates
        if name and name.lower().startswith("viraal"):
            continue
        x = ws.cell(row=r, column=2).value
        xii = ws.cell(row=r, column=3).value
        ug_score = ws.cell(row=r, column=5).value
        pg_score = ws.cell(row=r, column=8).value
        phd_score = ws.cell(row=r, column=11).value
        print(f"{str(name):<20} | {str(x):<8} | {str(xii):<9} | {str(ug_score or ''):<12} | {str(pg_score or ''):<12} | {str(phd_score or ''):<12}")
    wb.close()
    
    print("\n=== VERIFYING DETAILED EXPORT SCORES (VERCEL) ===")
    wb = openpyxl.load_workbook("vercel_detailed_export.xlsx")
    ws = wb.active
    print(f"{'Name':<20} | {'Class X':<8} | {'Class XII':<9} | {'Grad Score':<12} | {'PG Score':<12} | {'PhD Score':<12}")
    print("-" * 80)
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if name and name.lower().startswith("viraal"):
            continue
        x = ws.cell(row=r, column=12).value
        xii = ws.cell(row=r, column=13).value
        ug_score = ws.cell(row=r, column=16).value
        pg_score = ws.cell(row=r, column=20).value
        phd_score = ws.cell(row=r, column=24).value
        if name or x or xii or ug_score or pg_score or phd_score:
            print(f"{str(name or ''):<20} | {str(x or ''):<8} | {str(xii or ''):<9} | {str(ug_score or ''):<12} | {str(pg_score or ''):<12} | {str(phd_score or ''):<12}")
    wb.close()

if __name__ == '__main__':
    main()
