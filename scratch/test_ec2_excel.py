import requests
import openpyxl
from io import BytesIO

API = "http://13.205.216.81/api/v1"

# 1. Login
login_res = requests.post(f"{API}/auth/login", json={"username": "hr_ris", "password": "ris@1234"})
token = login_res.json().get("token")

# 2. Get Job ID
jobs = requests.get(f"{API}/public/jobs").json()
job_id = jobs[0]['id']
job_title = jobs[0]['title']

print(f"Testing Excel Export for Job: {job_title} ({job_id})")

# 3. Request Detailed Excel Export
res = requests.post(
    f"{API}/jobs/{job_id}/candidates/export",
    headers={"Authorization": f"Bearer {token}"},
    json={
        "filters": {},
        "format": "xlsx",
        "columns": [],
        "report_type": "detailed"
    }
)

print("Export Status Code:", res.status_code)
if res.status_code == 200:
    wb = openpyxl.load_workbook(BytesIO(res.content))
    sheet = wb.active
    headers = [cell.value for cell in sheet[1]]
    print(f"\nSUCCESS! Total Excel Columns Exported: {len(headers)}")
    print("Full Header Columns Breakdown:")
    for idx, h in enumerate(headers, 1):
        print(f" {idx:02d}. {h}")

    print(f"\nTotal Data Rows Exported: {sheet.max_row - 1}")
    if sheet.max_row > 1:
        row1_vals = [sheet.cell(row=2, column=col).value for col in range(1, len(headers)+1)]
        print("\nSample Row 1 Data Sample:")
        print(" Name:", row1_vals[0])
        print(" Email:", row1_vals[1])
        print(" Mobile:", row1_vals[3])
        print(" DOB:", row1_vals[4])
        print(" Age:", row1_vals[5])
        print(" SOP:", str(row1_vals[15])[:40] if row1_vals[15] else "N/A")
        print(" Class X:", row1_vals[27], row1_vals[28], row1_vals[29])
        print(" Class XII:", row1_vals[31], row1_vals[32], row1_vals[33])
        print(" Graduation:", row1_vals[35], row1_vals[36], row1_vals[37])
        print(" Total Exp:", row1_vals[55])
        print(" Last Salary:", row1_vals[56])
        print(" Work Org & Role:", row1_vals[57], row1_vals[58])
